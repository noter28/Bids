from fastapi import APIRouter, HTTPException

from app.api import crud
from app.api.models import BidSchema
from .auth import auth
from .excel_ops import write_excel, generate_excel_file
from typing import List
from openpyxl import Workbook, load_workbook
import os
import msal
import jwt
import json
import requests
from datetime import datetime

file_path = 'ОЕМ_  реестр_листопад.xlsx'
site_id = "8b2f1519-a497-4efe-8d83-b2a1d4d15d8e"
drive_id = 'b!GRUvi5ek_k6Ng7Kh1NFdjmu4BbllQgdHiWuZ2hmgRDm97BYuvkPkRLzzr8z5Gp3v'
folder_name = 'test2'
file_name = 'sample8.xlsx'
upload_url = f'https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}/items/root:/{folder_name}/{file_name}:/content'




router = APIRouter()


@router.post("/", status_code=201)
async def create_bid(payload: BidSchema):
    client_exist = await crud.client_exist(payload)
    if not client_exist:
        client_id = await crud.client_insert(payload)
    else:
        client_id = dict(client_exist._mapping)['id']
    bid_id = await crud.bid_insert(payload, client_id)
    response_object = {
        'status': 'success',
        'message': 'Bids for OSR was successfully added',
        'data': {
            'id': bid_id,
            'clientName': payload.client.name,
            'clientLeiCode': payload.client.leiCode,
            'clientType': payload.client.clientType,
            'beginDate': payload.beginDate,
            'osrName': payload.distributionCombinations[0]['osr']['name'],
            'osrLeiCode': payload.distributionCombinations[0]['osr']['leiCode'],
            'osrDistributions': payload.osrDistributions,
        }
    }
    return response_object


@router.post("/join/")
async def read_note(payload: List[BidSchema]):
    data = await crud.join(payload)
    return {
        'status': 'success',
        'message': 'Bids for OSR was successfully processed',
        'data': data
    }


@router.get("/excel/")
async def excel():

    # wb = load_workbook('document.xlsx')
    # wb.template = True
    # wb = Workbook()
    # ws = wb.active  # insert at the end (default)
    # ws['A5'] = 55
    # wb.save("sample2.xlsx")
    # return 'success'
    data = await crud.get_month_data('2023-08-01', 'АТ «ДТЕК ДНІПРОВСЬКІ ЕЛЕКТРОМЕРЕЖІ»')
    data = [dict(i._mapping) for i in data]
    wb = load_workbook('ОЕМ_  реестр_листопад.xlsx')
    wb.template = True
    wb.create_sheet('test')
    second_list = wb['test']
    list_of_lists = [[value for value in dictionary.values()] for dictionary in data]
    print(list_of_lists)
    for row in list_of_lists:
        second_list.append(row)
    wb.save("ОЕМ_  реестр_листопад.xlsx")
    return list_of_lists
    # accessToken = await auth()
    # write_excel(accessToken)
    # return await auth()


@router.post('/all_in_one/')
async def all(payload: List[BidSchema]):
    list_of_changed_DSO = await crud.is_exist(payload)
    for DSO in list_of_changed_DSO:  # list_of_changed_DSO = [[name1, date1], [name2, date2]]
        data = await crud.get_month_data(DSO['beginDate'], DSO['osr_name'])
        print(data)
        await generate_excel_file(data, DSO['link_to_template'])
    accessToken = await auth()
    await write_excel(accessToken, file_path, upload_url)
    return list_of_changed_DSO

@router.post("/exist/")
async def exist(payload: List[BidSchema]):
    a = await crud.is_exist(payload)
    return a




# @router.get("/{id}/", response_model=NoteDB)
# async def read_note(id: int):
#     note = await crud.get(id)
#     if not note:
#         raise HTTPException(status_code=404, detail="Note not found")
#     return note
#
#
# @router.get("/", response_model=List[NoteDB])
# async def read_all_notes():
#     return await crud.get_all()
#
#
# @router.put("/{id}/", response_model=NoteDB)
# async def update_note(id: int, payload: NoteSchema):
#     note = await crud.get(id)
#     if not note:
#         raise HTTPException(status_code=404, detail="Note not found")
#
#     note_id = await crud.put(id, payload)
#
#     response_object = {
#         "id": note_id,
#         "title": payload.title,
#         "description": payload.description,
#     }
#     return response_object
#
#
# @router.delete("/{id}/", response_model=NoteDB)
# async def delete_note(id: int):
#     note = await crud.get(id)
#     if not note:
#         raise HTTPException(status_code=404, detail="Note not found")
#
#     await crud.delete(id)
#
#     return note

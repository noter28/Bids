import requests
import os
from app.api import crud
from openpyxl import Workbook, load_workbook


file_path = 'sample2.xlsx'
site_id = "8b2f1519-a497-4efe-8d83-b2a1d4d15d8e"
drive_id = 'b!GRUvi5ek_k6Ng7Kh1NFdjmu4BbllQgdHiWuZ2hmgRDm97BYuvkPkRLzzr8z5Gp3v'
folder_name = 'test2'
file_name = 'sample8.xlsx'
upload_url = f'https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id}/items/root:/{folder_name}/{file_name}:/content'


async def generate_excel_file(data, file_name):
    # data = await crud.get_month_data('2023-08-01', 'АТ «ДТЕК ДНІПРОВСЬКІ ЕЛЕКТРОМЕРЕЖІ»')
    data = [dict(i._mapping) for i in data]
    wb = load_workbook(file_name)
    wb.template = True
    wb.create_sheet('test')
    second_list = wb['test']
    list_of_lists = [[value for value in dictionary.values()] for dictionary in data]
    print(1111111)
    print(list_of_lists)
    for row in list_of_lists:
        second_list.append(row)
    wb.save(file_name)
    return file_name



async def write_excel(token):
    try:
        with open(file_path, 'rb') as file:
            headers = {
                'Authorization': f"Bearer {token['access_token']}",
                'Content-Type': 'application/octet-stream',
                'Content-Length': str(os.path.getsize(file_path))
            }
            response = requests.put(upload_url, headers=headers, data=file)
    except Exception as err:
        print(err)
    return response.json()
